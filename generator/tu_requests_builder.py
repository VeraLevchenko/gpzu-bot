# generator/tu_requests_builder.py
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import List, Tuple, Dict, Optional
from datetime import date

from docxtpl import DocxTemplate
from openpyxl import load_workbook
from filelock import FileLock, Timeout

from parsers.egrn_parser import EGRNData


# Путь к корню проекта и папке с шаблонами
BASE_DIR = Path(__file__).resolve().parents[1]
TU_TEMPLATES_DIR = BASE_DIR / "templates" / "tu"

# Журнал
TU_JOURNAL_PATH = BASE_DIR / "Журнал запросов ТУ.xlsx"
TU_JOURNAL_LOCK_PATH = TU_JOURNAL_PATH.with_suffix(TU_JOURNAL_PATH.suffix + ".lock")

# Точное соответствие шаблонов и РСО:
#   suffix   – человекочитаемое для имени файла
#   rso_name – значение для столбца "РСО" в журнале
#   path     – путь к шаблону
TEMPLATE_CONFIG = [
    ("Водоканал", "водоснабжение", TU_TEMPLATES_DIR / "Водоканал.docx"),
    ("Газоснабжение", "газоснабжение", TU_TEMPLATES_DIR / "Газоснабжение.docx"),
    ("Теплоснабжение", "теплоснабжение", TU_TEMPLATES_DIR / "Теплоснабжение.docx"),
]


def _format_area(area: Optional[str]) -> str:
    """Аккуратное форматирование площади."""
    if not area:
        return ""
    s = area.strip().replace(",", ".")
    if s.endswith(".0"):
        s = s[:-2]
    return s


def build_tu_context(
    egrn: EGRNData,
    incoming: str,
    out_num: Optional[str] = None,
    out_date: Optional[str] = None,
) -> Dict[str, str]:
    """
    Контекст для подстановки в шаблоны.

    Шаблоны используют плейсхолдеры:
      {{INCOMING}}, {{CADNUM}}, {{AREA}}, {{VRI}}, {{ADDRESS}},
      {{OUT_NUM}}, {{OUT_DATE}}
    """
    return {
        "INCOMING": incoming or "",
        "CADNUM": egrn.cadnum or "",
        "AREA": _format_area(egrn.area),
        "VRI": egrn.permitted_use or "",
        "ADDRESS": egrn.address or "",
        "OUT_NUM": out_num or "",
        "OUT_DATE": out_date or "",
    }


def _render_doc(template_path: Path, context: Dict[str, str]) -> bytes:
    """Рендер DOCX по шаблону."""
    tpl = DocxTemplate(str(template_path))
    tpl.render(context)
    bio = BytesIO()
    tpl.save(bio)
    return bio.getvalue()


def build_tu_docs(
    egrn: EGRNData,
    incoming: str,
    out_num: Optional[str] = None,
    out_date: Optional[str] = None,
) -> List[Tuple[str, bytes]]:
    """
    Вариант без исходящего номера: формируем документы по шаблонам,
    но поля OUT_NUM и OUT_DATE остаются пустыми (или заданными явно).
    Журнал НЕ трогаем.
    """
    ctx = build_tu_context(egrn, incoming, out_num=out_num, out_date=out_date)
    docs: List[Tuple[str, bytes]] = []

    cad = (egrn.cadnum or "no_cad").replace(":", " ")

    for suffix, _rso_name, tpl_path in TEMPLATE_CONFIG:
        if not tpl_path.exists():
            continue

        content = _render_doc(tpl_path, ctx)
        filename = f"ТУ_{suffix}_{cad}.docx"
        docs.append((filename, content))

    return docs


def build_tu_docs_with_outgoing(egrn: EGRNData, incoming: str) -> List[Tuple[str, bytes]]:
    """
    Вариант с исходящими:

      - Берём файловую блокировку журнала (FileLock) с таймаутом,
        чтобы только один процесс в данный момент работал с Excel.
      - Внутри блокировки:
          * открываем журнал
          * находим максимальный исходящий номер
          * для КАЖДОГО шаблона:
              - номер = текущий + 1
              - добавляем строку в журнал с РСО
              - формируем DOCX с этим номером и сегодняшней датой
          * сохраняем журнал один раз.

    Если журнал занят (кто-то другой в этот момент формирует ТУ) —
    выбрасывается понятная ошибка.

    Если Excel-файл открыт в другом приложении и не даёт сохранить —
    также выбрасывается человекочитаемая ошибка.
    """
    if not TU_JOURNAL_PATH.exists():
        raise FileNotFoundError(f"Не найден журнал ТУ: {TU_JOURNAL_PATH}")

    # таймаут на блокировку — 3 секунды
    lock = FileLock(str(TU_JOURNAL_LOCK_PATH), timeout=3)

    try:
        with lock:
            try:
                wb = load_workbook(TU_JOURNAL_PATH)
            except PermissionError:
                # если даже прочитать не можем — значит файл плотно занят
                raise RuntimeError(
                    "Не удалось открыть журнал запросов ТУ. "
                    "Возможно, файл открыт другим пользователем. "
                    "Закройте журнал и попробуйте ещё раз."
                )

            ws = wb.active  # по умолчанию Лист1

            # первая строка — заголовки
            headers = {cell.value: cell.column for cell in ws[1]}

            col_out_num = headers.get("Исходящий номер")
            col_out_date = headers.get("Исходящая дата")
            col_incoming = headers.get("Номер Заявления и дата")
            col_cadnum = headers.get("Кадастровый номер земельного участка")
            col_addr = headers.get("Адрес")
            col_rso = headers.get("РСО")

            if not all([col_out_num, col_out_date, col_incoming, col_cadnum, col_addr, col_rso]):
                raise RuntimeError(
                    "В журнале ТУ отсутствуют необходимые столбцы. "
                    "Ожидаются: 'Исходящий номер', 'Исходящая дата', "
                    "'Номер Заявления и дата', 'Кадастровый номер земельного участка', "
                    "'Адрес', 'РСО'."
                )

            # Находим максимальный существующий исходящий номер
            max_num = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col_out_num).value
                if val is None:
                    continue
                try:
                    n = int(str(val).strip())
                except Exception:
                    continue
                if n > max_num:
                    max_num = n

            current_num = max_num
            today_str = date.today().strftime("%d.%m.%Y")

            docs: List[Tuple[str, bytes]] = []
            cad = (egrn.cadnum or "no_cad").replace(":", " ")

            # Для каждого шаблона — свой номер и своя строка в журнале
            for suffix, rso_name, tpl_path in TEMPLATE_CONFIG:
                if not tpl_path.exists():
                    continue

                # следующий исходящий номер
                current_num += 1
                out_num_str = str(current_num)
                out_date_str = today_str

                # записываем строку в журнал
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=col_out_num, value=current_num)
                ws.cell(row=new_row, column=col_out_date, value=out_date_str)
                ws.cell(row=new_row, column=col_incoming, value=incoming)
                ws.cell(row=new_row, column=col_cadnum, value=egrn.cadnum or "")
                ws.cell(row=new_row, column=col_addr, value=egrn.address or "")
                ws.cell(row=new_row, column=col_rso, value=rso_name)

                # формируем DOCX с этим номером и датой
                ctx = build_tu_context(egrn, incoming, out_num=out_num_str, out_date=out_date_str)
                content = _render_doc(tpl_path, ctx)
                filename = f"ТУ_{suffix}_{cad}.docx"
                docs.append((filename, content))

            # сохраняем журнал один раз в конце
            try:
                wb.save(TU_JOURNAL_PATH)
            except PermissionError:
                # классическая ситуация: Excel/LibreOffice держит файл на записи
                raise RuntimeError(
                    "Не удалось сохранить журнал запросов ТУ. "
                    "Возможно, файл открыт другим пользователем. "
                    "Закройте журнал и повторите попытку."
                )
            except OSError as ex:
                # прочие ошибки файловой системы — тоже объясняем по-человечески
                raise RuntimeError(
                    f"Ошибка при сохранении журнала запросов ТУ: {ex}"
                )

    except Timeout:
        # если за 3 секунды не удалось захватить lock-файл
        raise RuntimeError(
            "Журнал запросов ТУ сейчас используется другим процессом. "
            "Попробуйте выполнить формирование ТУ немного позже."
        )

    return docs
